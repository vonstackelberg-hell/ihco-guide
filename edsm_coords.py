#!/usr/bin/env python3
"""
EDSM Koordinaten-Abfrage für Trading.xlsx
Füllt Spalten D (SystemID/URL) und E-G (X/Y/Z) mit Daten von EDSM.
"""

import requests
import time
from openpyxl import load_workbook
from openpyxl.styles import Font

XLSX_PATH = "Trading.xlsx"
OUTPUT_PATH = "Trading_coords.xlsx"
EDSM_API = "https://www.edsm.net/api-v1/systems"

def get_coords(system_names):
    """Holt Koordinaten für bis zu 100 Systeme auf einmal."""
    params = {"showCoordinates": 1, "showId": 1}
    for i, name in enumerate(system_names):
        params[f"systemName[{i}]"] = name
    try:
        r = requests.get(EDSM_API, params=params, timeout=15)
        r.raise_for_status()
        return {s["name"].lower(): s for s in r.json()}
    except Exception as e:
        print(f"  Fehler: {e}")
        return {}

def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    # Header ergänzen falls nötig
    ws["E1"] = "X"
    ws["F1"] = "Y"
    ws["G1"] = "Z"
    ws["D1"] = "EDSM URL"
    for col in ["D", "E", "F", "G"]:
        ws[f"{col}1"].font = Font(bold=True)

    # Alle Systeme aus Spalte C sammeln (unique)
    systems = {}
    for row in ws.iter_rows(min_row=2):
        system = row[2].value  # Spalte C
        if system and system.strip():
            systems.setdefault(system.strip(), []).append(row[0].row)

    system_names = list(systems.keys())
    print(f"Systeme gefunden: {len(system_names)}")

    # In Batches à 100 abfragen
    results = {}
    for i in range(0, len(system_names), 100):
        batch = system_names[i:i+100]
        print(f"Abfrage Batch {i//100 + 1} ({len(batch)} Systeme)...")
        results.update(get_coords(batch))
        time.sleep(0.5)

    # Koordinaten in Excel eintragen
    found = 0
    not_found = []
    for system, rows in systems.items():
        data = results.get(system.lower())
        if data:
            found += 1
            coords = data.get("coords", {})
            edsm_id = data.get("id", "")
            edsm_url = f"https://www.edsm.net/en/system/id/{edsm_id}" if edsm_id else ""
            for row_num in rows:
                ws.cell(row=row_num, column=4).value = edsm_url
                ws.cell(row=row_num, column=5).value = coords.get("x")
                ws.cell(row=row_num, column=6).value = coords.get("y")
                ws.cell(row=row_num, column=7).value = coords.get("z")
        else:
            not_found.append(system)

    wb.save(OUTPUT_PATH)
    print(f"\n✓ Fertig! {found}/{len(system_names)} Systeme gefunden.")
    if not_found:
        print(f"Nicht gefunden ({len(not_found)}):")
        for s in not_found:
            print(f"  - {s}")
    print(f"\nGespeichert als: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
